from customtkinter import *
import os
from tkinter import *
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

import keyboard
import pyautogui
import time

set_appearance_mode("dark") 
set_default_color_theme("blue")

t = CTk()
t.geometry("800x500")
t.resizable(False, False)

texto = CTkLabel(t, width=200, text="Selecione a planilha", font=("Arial", 20, "bold"))
texto.pack(pady=20)

f = [i for i in os.listdir() if i.endswith(".xlsx")]

def adicionar():
	global cpf_column, nome_column, sheet, add_button
	texto.configure(text="bagulho aqui")
	s = sheet.cell(1, cpf_column + 1).value
	if "str" in str(type(s)):
		sheet.delete_rows(1)
	
	time.sleep(10)
	for i in sheet.values:
		pyautogui.write(str(i[cpf_column]))
		pyautogui.press("tab")
		pyautogui.write(str(i[nome_column]))
		pyautogui.press("tab")
		pyautogui.press("tab")
		if keyboard.is_pressed("esc"):
			break
	try:
		for i in range(0, 3):
		    pyautogui.press("tab")
	finally:
		pyautogui.press("enter")
	
def deletar():
	global cpf_column, nome_column, sheet, del_button
	texto.configure(text="bagulho aqui")
	cell = sheet.cell(1, cpf_column + 1).value
	if "str" in str(type(cell)):
		sheet.delete_rows(1)
	
	time.sleep(10)
	for valor in sheet.values:
		pyautogui.write(str(valor[cpf_column]))
		pyautogui.press("tab")
		pyautogui.write(str(valor[nome_column]))
		pyautogui.press("tab")
		pyautogui.press("tab")
		if keyboard.is_pressed("esc"):
			break
	try:
		for i in range(0, 4):
		    pyautogui.press("tab")
	finally:
		pyautogui.press("enter")
	 
def get_cpf_column(column):
	global cpf_column, nome_column, add_button, del_button
	cpf_column = column_index_from_string(column) - 1

	cb.destroy()

	texto.configure(text="Tudo certinho! Só apertar o botão abaixo para iniciar o processo.")
	
	add_button = CTkButton(t, text="Adicionar", width=100, command=adicionar).pack(side="left", padx=(290, 10), pady=(0, 400))
	del_button = CTkButton(t, text="Deletar", width=100, command=deletar).pack(side="left", padx=(10, 290), pady=(0, 400))

def get_nome_column(column):
	global colunas, nome_column
	nome_column = column_index_from_string(column) - 1
	texto.configure(text="Selecione a coluna em que está o CPF")
	
	colunas = [x for x in colunas if x != column]
	
	cb.configure(values=colunas, command=get_cpf_column)
	cb.set("--selecione--")

def planilha(_):
    global colunas, sheet
    planilha = cb.get()
    wb = load_workbook(planilha)
    sheet = wb.active
    
    max_column = sheet.max_column
    
    colunas = [get_column_letter(i) for i in range(1, max_column + 1)]
    
    texto.configure(text="Selecione a coluna em que está o nome")

    cb.configure(values=colunas, command=get_nome_column)
    cb.set("--selecione--")
    
cb = CTkComboBox(t, width=300, values=f, command=planilha)
cb.pack()
cb.set("--selecione--")
t.mainloop()
